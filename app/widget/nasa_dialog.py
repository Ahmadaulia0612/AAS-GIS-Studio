import os
import requests
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

from PySide6.QtCore import QDate, QObject, QThread, Signal, Slot
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QLabel, QComboBox, QDateEdit,
    QPushButton, QMessageBox, QFileDialog, QHBoxLayout, QWidget
)


class CHIRPSDownloadWorker(QObject):
    """
    Worker khusus CHIRPS.

    Tujuannya hanya memindahkan request Climate Engine dan pengolahan
    DataFrame ke background thread supaya GUI tidak menjadi Not Responding.
    Alur hasil dan format Excel tetap sama.
    """

    finished = Signal(object, object)
    error = Signal(str)

    def __init__(self, lat, lon, start_date, end_date):
        super().__init__()
        self.lat = lat
        self.lon = lon
        self.start_date = start_date
        self.end_date = end_date

    @Slot()
    def run(self):
        try:
            df_daily, monthly = self.download()
            self.finished.emit(df_daily, monthly)
        except Exception as e:
            self.error.emit(str(e))

    def download(self):
        url = "https://api.climateengine.org/timeseries/native/coordinates"

        payload = {
            "coordinates": f"[[{float(self.lon)}, {float(self.lat)}]]",
            "area_reducer": "mean",
            "dataset": "CHIRPS_DAILY",
            "variable": "precipitation",
            "start_date": self.start_date,
            "end_date": self.end_date
        }

        # Climate Engine API key dibaca dari environment.
        api_key = os.environ.get("CLIMATE_ENGINE_API_KEY", "").strip()

        if not api_key:
            raise Exception(
                "CLIMATE_ENGINE_API_KEY belum tersedia. "
                "Set environment variable terlebih dahulu."
            )

        # Endpoint Climate Engine menerima token langsung.
        headers = {
            "Authorization": api_key,
            "Content-Type": "application/json"
        }

        response = requests.post(
            url,
            json=payload,
            headers=headers,
            timeout=120
        )

        if response.status_code != 200:
            raise Exception(
                f"Gagal API Climate Engine ({response.status_code}): "
                f"{response.text}"
            )

        result = response.json()

        try:
            raw_data = result["Data"][0]["Data"]
        except (KeyError, IndexError, TypeError):
            raise Exception(
                "Respons Climate Engine tidak memiliki struktur Data yang diharapkan."
            )

        df = pd.DataFrame(raw_data)

        # Climate Engine mengembalikan nama kolom:
        # "Date" dan "precipitation (mm)"
        df.rename(
            columns={
                "Date": "Periode",
                "precipitation (mm)": "Precipitation (mm)"
            },
            inplace=True
        )

        required_cols = {"Periode", "Precipitation (mm)"}
        missing_cols = required_cols - set(df.columns)

        if missing_cols:
            raise Exception(
                "Kolom CHIRPS tidak sesuai respons API. "
                f"Kolom hilang: {sorted(missing_cols)}"
            )

        df["Periode"] = pd.to_datetime(
            df["Periode"],
            errors="coerce"
        )

        df["Precipitation (mm)"] = pd.to_numeric(
            df["Precipitation (mm)"],
            errors="coerce"
        ).fillna(0.0)

        df = df.dropna(subset=["Periode"])

        # Rekap Bulanan
        df_proc = df.copy()
        df_proc["YEAR"] = df_proc["Periode"].dt.year
        df_proc["MONTH"] = df_proc["Periode"].dt.month

        monthly = (
            df_proc.pivot_table(
                index="YEAR",
                columns="MONTH",
                values="Precipitation (mm)",
                aggfunc="sum",
                fill_value=0
            )
            .reset_index()
        )

        for month in range(1, 13):
            if month not in monthly.columns:
                monthly[month] = 0

        monthly = monthly[
            ["YEAR"] + list(range(1, 13))
        ]

        monthly.columns = [
            "YEAR",
            "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
            "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"
        ]

        monthly["ANN"] = monthly[
            [
                "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"
            ]
        ].sum(axis=1)

        # Format harian
        df_daily = df[
            ["Periode", "Precipitation (mm)"]
        ].copy()

        df_daily["Periode"] = df_daily[
            "Periode"
        ].dt.strftime("%Y-%m-%d")

        return df_daily, monthly


class ClimateDataDownloadDialog(QDialog):
    def __init__(self, lat, lon, parent=None):
        super().__init__(parent)

        self.lat = lat
        self.lon = lon

        self.setWindowTitle("Unduh Data Iklim Presisi (NASA & CHIRPS)")
        self.setFixedSize(450, 400)

        # Thread CHIRPS disimpan sebagai atribut supaya tidak dihancurkan
        # sebelum proses background selesai.
        self.chirps_thread = None
        self.chirps_worker = None

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)

        # Info Koordinat
        self.lbl_info = QLabel(
            f"<b>Koordinat Target:</b><br>"
            f"Latitude: {self.lat:.4f}, Longitude: {self.lon:.4f}"
        )
        layout.addWidget(self.lbl_info)

        # Pilihan Sumber Data
        layout.addWidget(QLabel("Pilih Sumber Data:"))

        self.combo_source = QComboBox()
        self.combo_source.addItems([
            "NASA POWER (Unduh Online)",
            "CHIRPS Asli (Climate Engine API)"
        ])
        self.combo_source.currentIndexChanged.connect(
            self.on_source_changed
        )
        layout.addWidget(self.combo_source)

        # Resolusi Waktu NASA
        self.lbl_interval = QLabel("Pilih Resolusi Waktu:")
        layout.addWidget(self.lbl_interval)

        self.combo_interval = QComboBox()
        self.combo_interval.addItems([
            "Harian (Daily)",
            "Bulanan (Monthly)",
            "Tahunan (Annual)"
        ])
        layout.addWidget(self.combo_interval)

        # Rentang tanggal
        date_widget_container = QWidget()
        date_layout = QHBoxLayout(date_widget_container)
        date_layout.setContentsMargins(0, 0, 0, 0)

        self.date_start = QDateEdit()
        self.date_start.setCalendarPopup(True)
        self.date_start.setDate(QDate(2010, 1, 1))

        self.date_end = QDateEdit()
        self.date_end.setCalendarPopup(True)
        self.date_end.setDate(QDate(2025, 12, 31))

        v1 = QVBoxLayout()
        v1.addWidget(QLabel("Tanggal Mulai:"))
        v1.addWidget(self.date_start)

        v2 = QVBoxLayout()
        v2.addWidget(QLabel("Tanggal Selesai:"))
        v2.addWidget(self.date_end)

        date_layout.addLayout(v1)
        date_layout.addLayout(v2)

        layout.addWidget(date_widget_container)

        # Tombol aksi
        self.btn_download = QPushButton(
            "📥 Unduh Otomatis & Simpan (.xlsx)"
        )
        self.btn_download.setStyleSheet(
            "background-color: #28a745; "
            "color: white; "
            "font-weight: bold; "
            "padding: 10px;"
        )
        self.btn_download.clicked.connect(self.process_data)

        layout.addWidget(self.btn_download)

    def on_source_changed(self, index):
        if index == 0:
            self.lbl_interval.show()
            self.combo_interval.show()
        else:
            # CHIRPS Climate Engine secara default harian.
            self.lbl_interval.hide()
            self.combo_interval.hide()

    def save_chirps_excel(self, filepath, df_daily, monthly):
        with pd.ExcelWriter(
            filepath,
            engine="openpyxl"
        ) as writer:
            df_daily.to_excel(
                writer,
                sheet_name="Data CHIRPS Asli",
                index=False,
                startrow=0,
                startcol=0
            )

            monthly.to_excel(
                writer,
                sheet_name="Data CHIRPS Asli",
                index=False,
                startrow=1,
                startcol=4
            )

        wb = load_workbook(filepath)
        ws = wb["Data CHIRPS Asli"]

        ws["E1"] = "HASIL REKAPITULASI BULAN"

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 22

        for col in range(5, 19):
            ws.column_dimensions[
                chr(64 + col)
            ].width = 12

        wb.save(filepath)

    def process_data(self):
        source_idx = self.combo_source.currentIndex()

        start_str = self.date_start.date().toString(
            "yyyy-MM-dd"
        )
        end_str = self.date_end.date().toString(
            "yyyy-MM-dd"
        )

        # Validasi tanggal
        if self.date_start.date() > self.date_end.date():
            QMessageBox.warning(
                self,
                "Tanggal Tidak Valid",
                "Tanggal mulai tidak boleh lebih besar "
                "dari tanggal selesai."
            )
            return

        self.btn_download.setEnabled(False)
        self.btn_download.setText(
            "Sedang memproses data..."
        )

        if source_idx == 1:
            # ==========================================================
            # CHIRPS
            # ==========================================================
            # HANYA bagian ini yang dipindahkan ke background thread.
            # Request dan pengolahan DataFrame tidak lagi memblokir GUI.
            self.btn_download.setText(
                "Sedang mengunduh CHIRPS..."
            )

            self.chirps_thread = QThread(self)
            self.chirps_worker = CHIRPSDownloadWorker(
                self.lat,
                self.lon,
                start_str,
                end_str
            )

            self.chirps_worker.moveToThread(
                self.chirps_thread
            )

            self.chirps_thread.started.connect(
                self.chirps_worker.run
            )

            self.chirps_worker.finished.connect(
                self.on_chirps_finished
            )

            self.chirps_worker.error.connect(
                self.on_chirps_error
            )

            # Setelah worker selesai, thread dihentikan.
            self.chirps_worker.finished.connect(
                self.chirps_thread.quit
            )
            self.chirps_worker.error.connect(
                self.chirps_thread.quit
            )

            self.chirps_thread.finished.connect(
                self.on_chirps_thread_finished
            )

            self.chirps_thread.start()
            return

        # ==============================================================
        # NASA POWER
        # ==============================================================
        # Jalur NASA tetap seperti sebelumnya.
        try:
            interval = [
                "daily",
                "monthly",
                "annual"
            ][self.combo_interval.currentIndex()]

            api_interval = (
                "daily"
                if interval == "daily"
                else interval
            )

            s_nasa = (
                self.date_start.date().toString("yyyyMMdd")
                if api_interval == "daily"
                else self.date_start.date().toString("yyyy")
            )

            e_nasa = (
                self.date_end.date().toString("yyyyMMdd")
                if api_interval == "daily"
                else self.date_end.date().toString("yyyy")
            )

            url = (
                "https://power.larc.nasa.gov/api/"
                f"temporal/{api_interval}/point"
            )

            params = {
                "parameters": "PRECTOTCORR",
                "community": "ag",
                "longitude": self.lon,
                "latitude": self.lat,
                "start": s_nasa,
                "end": e_nasa,
                "format": "JSON"
            }

            res = requests.get(
                url,
                params=params,
                timeout=35
            ).json()

            props = res[
                "properties"
            ][
                "parameter"
            ][
                "PRECTOTCORR"
            ]

            df = pd.DataFrame(
                list(props.items()),
                columns=[
                    "Periode",
                    "Precipitation (mm)"
                ]
            )

            filepath, _ = QFileDialog.getSaveFileName(
                self,
                "Simpan File",
                "nasa_rainfall_combined.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not filepath:
                self.reset_button()
                return

            with pd.ExcelWriter(
                filepath,
                engine="openpyxl"
            ) as writer:

                df_out = df.copy()

                dt_series = pd.to_datetime(
                    df_out["Periode"],
                    format="%Y%m%d",
                    errors="coerce"
                )

                if interval == "daily":
                    df_out["Periode"] = (
                        dt_series.dt.strftime("%Y-%m-%d")
                    )

                df_out.to_excel(
                    writer,
                    sheet_name="Data NASA",
                    index=False
                )

                if interval == "daily":
                    df_p = df.copy()

                    df_p["Year"] = dt_series.dt.year
                    df_p["Month"] = dt_series.dt.month

                    pivot = df_p.pivot_table(
                        index="Year",
                        columns="Month",
                        values="Precipitation (mm)",
                        aggfunc="sum"
                    )

                    pivot.columns = [
                        "JAN", "FEB", "MAR", "APR",
                        "MAY", "JUN", "JUL", "AUG",
                        "SEP", "OCT", "NOV", "DEC"
                    ]

                    pivot["ANN"] = pivot.sum(axis=1)

                    pivot.to_excel(
                        writer,
                        sheet_name="Data NASA",
                        startrow=0,
                        startcol=4
                    )

            QMessageBox.information(
                self,
                "Berhasil",
                f"Data NASA berhasil disimpan di:\n{filepath}"
            )

            self.accept()

        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                str(e)
            )

        finally:
            self.reset_button()

    @Slot(object, object)
    def on_chirps_finished(self, df_daily, monthly):
        """
        Dipanggil di GUI thread setelah download CHIRPS selesai.

        QFileDialog memang sengaja tetap di GUI thread.
        """
        try:
            self.btn_download.setText(
                "Memilih lokasi penyimpanan..."
            )

            filepath, _ = QFileDialog.getSaveFileName(
                self,
                "Simpan File",
                "chirps_rekap_climate_engine.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not filepath:
                self.reset_button()
                return

            self.btn_download.setText(
                "Menyimpan Excel..."
            )

            self.save_chirps_excel(
                filepath,
                df_daily,
                monthly
            )

            QMessageBox.information(
                self,
                "Berhasil",
                "Data CHIRPS berhasil diunduh dan "
                f"direkap ke:\n{filepath}"
            )

            self.accept()

        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                str(e)
            )

        finally:
            self.reset_button()

    @Slot(str)
    def on_chirps_error(self, err_msg):
        QMessageBox.critical(
            self,
            "Error CHIRPS",
            err_msg
        )

        self.reset_button()

    @Slot()
    def on_chirps_thread_finished(self):
        # Worker/thread hanya dibersihkan setelah QThread benar-benar selesai.
        if self.chirps_worker is not None:
            self.chirps_worker.deleteLater()

        if self.chirps_thread is not None:
            self.chirps_thread.deleteLater()

        self.chirps_worker = None
        self.chirps_thread = None

    def reset_button(self):
        self.btn_download.setEnabled(True)
        self.btn_download.setText(
            "📥 Unduh Otomatis & Simpan (.xlsx)"
        )

    def closeEvent(self, event):
        # Jangan menutup dialog ketika worker masih berjalan.
        if (
            self.chirps_thread is not None
            and self.chirps_thread.isRunning()
        ):
            QMessageBox.information(
                self,
                "Proses Masih Berjalan",
                "Download CHIRPS masih berjalan. "
                "Tunggu sampai selesai."
            )
            event.ignore()
            return

        event.accept()